from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.utils import translation

from .models import *
import requests
from .forms import ContactForm, MxikForm

import pandas as pd
import openpyxl
import requests
from django.shortcuts import redirect

# Create your views here.
token = '5782879949:AAFamiP8CUietgluk-LSn4zLAe23qLHImkQ'
group_chat_id = -719945505

check_mxik_code_url = "https://api-tasnif.soliq.uz/cl-api/integration-mxik/get/history/"
result_list = []


def send_message(user_id, text):
    api_endpoint = requests.get(
        'https://api.telegram.org/bot{}/sendMessage?chat_id={}&text={}&parse_mode=HTML'.format(token, user_id, text))


def index(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)

        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            text = form.cleaned_data['text']
            whole_text = f"<b>Имя:</b> {name}\n<b>Эл. адрес:</b> {email}\n<b>Текст:</> {text}"
            send_message(group_chat_id, whole_text)
        else:
            return render(request, "index-11.html", {'form': form})
    try:
        oferta_umumiy = CompanyDocument.objects.get(file_name="umumiy oferta")
        oferta_snq = CompanyDocument.objects.get(file_name="snq oferta")
    except:
        oferta_umumiy = "#"
        oferta_snq = "#"

    form = ContactForm()

    return render(request, "index-11.html", {'form': form, "oferta_umumiy": oferta_umumiy, "oferta_snq": oferta_snq})


def blog(request):
    blogs = Blog.objects.all()

    context = {"blogs": blogs}
    return render(request, "blog.html", context)


def handle_uploaded_file(f):
    with open('geeks / upload/'+f.name, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def check_mxik(request):
    if request.method == 'POST':
        form = MxikForm(request.POST, request.FILES)
        if form.is_valid():
            mxik_object_id = form.save_and_get_id()
            continue_proc = check_the_website()
            if continue_proc:
                request_to_website(mxik_object_id)

            mxikobj = MxikObject.objects.get(id=mxik_object_id)

            context = {"form": form, "output": mxikobj}
            return render(request, "form.html", context)
    else:
        form = MxikForm()
    context = {"form": form}
    return render(request, "form.html", context)


def server_down(request):
    form = MxikForm()
    message = "Tasnif serveri ishlamayapti"
    context = {"form": form, "message": message}
    return render(request, "blog.html", context)


def blog_details(request, id):
    blog = Blog.objects.get(id=id)
    blogs = Blog.objects.filter(popular=True)

    context = {"blog": blog, "blogs": blogs}
    return render(request, "blog-details.html", context)


def check_the_website() -> bool:
    response = requests.get(check_mxik_code_url + "02199001005000000")
    if response.status_code == 200:
        return True
    return False


def request_to_website(file_id):
    try:
        mxik_model = MxikObject.objects.get(id=file_id)
        wb_obj = openpyxl.load_workbook(mxik_model.mxik_input.name)
        sheet_obj = wb_obj.active
        count_row = sheet_obj.max_row
        count_column = sheet_obj.max_column
        print(count_row)
        if mxik_model.mxik_options.option == "DEVICE":
            column_index_mxik = 8
        else:
            column_index_mxik = 11

        for i in range(1, count_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=column_index_mxik)
            mxik_code = cell_obj.value
            if mxik_code:
                single_column_from_excel = get_column_from_excel(
                    sheet_obj, i)
                respond = check_mxik_from_tasnif(mxik_code)

                check_result = parse_data(
                    respond, single_column_from_excel, mxik_model.mxik_options.option)
                print(i, respond, check_result)
                if check_result == False:
                    return False

        output_name = mxik_model.mxik_input.name.replace(
            '/mxik', '/output')
        pd.DataFrame(result_list).to_excel(
            output_name, header=False, index=False)
        mxik_model.mxik_output = output_name
        mxik_model.save()
        return True

    except requests.exceptions.RequestException as e:
        # Handle any other exception
        return False


def check_mxik_from_tasnif(mxik_code):
    mxik_code_res = requests.get(check_mxik_code_url + str(mxik_code))
    return mxik_code_res


def parse_data(data, column_excel, mxik_type):
    data = data.json()

    if data['data']:
        mxik_status = data['data']['status']
        mxik_code = data['data']['mxikCode']
        update_excel(column=column_excel,
                     mxik_status=mxik_status, mxik_code=mxik_code, mxik_type=mxik_type)
        return True
    return True


def update_excel(column, mxik_status, mxik_code, mxik_type):
    # mxid_code does not exists
    if mxik_status == -2:
        print("deleting the row -> ", mxik_status, mxik_code)

    # deleted mxik_code
    elif mxik_status == -1:
        print("deleting the row -> ", mxik_status, mxik_code)

    # editing changed mxik_code
    elif mxik_status == 2:
        if mxik_type == "DEVICE":
            column[-1] = mxik_code
        else:
            column[-2] = mxik_code
        result_list.append(column)

    # already fine mxik_code
    else:
        result_list.append(column)


def get_column_from_excel(sheet, row_num):
    column = []
    for i in range(1, 9):
        single_column = sheet.cell(row=row_num, column=i)
        column.append(single_column.value)
    return column


def change_lang(request):
    LANGUAGE_SESSION_KEY = '_language'
    if request.method == "POST":
        sent_url = request.POST['next']
        old_lang = request.LANGUAGE_CODE
        changed_lang = request.POST['language']
        translation.activate(changed_lang)
        request.session[LANGUAGE_SESSION_KEY] = changed_lang
        # I use HTTP_REFERER to direct them back to previous path
        print(sent_url)
        if "en" in sent_url:
            if changed_lang != 'uz':
                new_url = sent_url.replace('en', changed_lang)
                print(new_url)
                return HttpResponseRedirect(new_url)
            elif changed_lang == 'uz':
                new_url1 = sent_url.replace('en', '')
                new_url = new_url1[1:]
                print(new_url)
                return HttpResponseRedirect(new_url)
        elif "ru" in sent_url:
            if changed_lang != 'uz':
                new_url = sent_url.replace('ru', changed_lang)
                print(new_url)
                return HttpResponseRedirect(new_url)
            elif changed_lang == 'uz':
                new_url1 = sent_url.replace('ru', '')
                new_url = new_url1[1:]
                print(new_url)
                return HttpResponseRedirect(new_url)
        elif old_lang == "uz" and changed_lang != 'uz':
            new_url = f"/{changed_lang}" + sent_url

            return HttpResponseRedirect(new_url)

        return HttpResponseRedirect(sent_url)
